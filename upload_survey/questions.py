from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import sqlite3
import sys

allowed_q_types = ["evaluation", "choosing"]    # Доступные типы вопросов


def parse_table(table):
    questions = list()
    index_cell_j = 2
    while table.cell(1, index_cell_j).value is not None:
        questions.append({'question': table.cell(1, index_cell_j).value, \
            'type': table.cell(2, index_cell_j).value, 'options': list()})
        index_cell_i = 3
        while table.cell(index_cell_i, index_cell_j).value is not None:
            questions[-1]['options'].append(table.cell(index_cell_i, index_cell_j).value)
            index_cell_i += 1
        index_cell_j += 1
    return (questions)


def check_questions(list_q: list):
    for q in list_q:
        if q["type"] not in allowed_q_types:
            exit(2)     # error QuestionTypes
        if (len(q["options"]) < 2 and q["type"] == "choosing") or \
            (len(q["options"]) != 1 and q["type"] == "evaluation"):
            exit(3)     # error NumberOfOptions


def load_question_types(cursor):
    id_l = list()
    for type in allowed_q_types:
        id = cursor.execute("SELECT id FROM QuestionTypes WHERE type == ?", (type,)).fetchone()
        if id is None:
            cursor.execute("INSERT INTO QuestionTypes (type) VALUES(?)", (type,))
            id = cursor.lastrowid
        else:
            id = id[0]
        id_l.append(id)
    return id_l


def load_survey(cursor):
    cursor.execute("INSERT INTO Surveys (file_id) VALUES(?)", (None,))
    return cursor.lastrowid


def get_id_type(type: str, ids: list):
    for i in range(0, len(allowed_q_types)):
        if allowed_q_types[i] == type:
            return ids[i]


def load_questions(cursor, id_s, id_t_l, list_q):
    id_q_l = list()
    for i in range(0, len(list_q)):
        question = list_q[i]
        cursor.execute("INSERT INTO Questions (id_survey, question, id_type) VALUES(?, ?, ?)", \
                        (id_s, question["question"], get_id_type(question["type"], id_t_l)))
        id_q_l.append(cursor.lastrowid)
    return id_q_l


def load_options(cursor, list_q, id_q_l):
    for i in range(0, len(list_q)):
        if list_q[i]["type"] == "evaluation":
            list_q[i]["options"] = [i for i in range(1, 1 + list_q[i]["options"][0])]   # creating range of rates for eval type
        cursor.executemany("INSERT INTO Options (option, id_question) VALUES(?, ?)", \
                            tuple((q, id_q_l[i]) for q in list_q[i]["options"]))


def load_in_db(list_q: list, table: Worksheet):
    try:
        conn = sqlite3.connect("../db/database.db")
    except (sqlite3.OperationalError):
        exit(4)     # Can't open/create file "../db/database.db"
    cursor = conn.cursor()  
    with open("../db/create_database.sql", "r") as f:
        sql = f.read()
    cursor.executescript(sql)

    id_t_l = load_question_types(cursor)    # ret: ids of types in DB
    id_s = load_survey(cursor)              # ret: id of a new servey
    id_q_l = load_questions(cursor, id_s, id_t_l, list_q)   # ret: ids of loaded questions
    load_options(cursor, list_q, id_q_l)

    conn.commit()
    conn.close()


def create_new_survey(file):
    wb = load_workbook(filename = file)
    table = wb['Survey']
    question_list = parse_table(table)
    check_questions(question_list)
    load_in_db(question_list, table)


if len(sys.argv) == 1:
    pathfile = 'questions.xlsx'
else:
    pathfile = sys.argv[1]

create_new_survey(pathfile)